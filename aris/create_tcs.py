from PyQt5 import QtCore
import pandas as pd
import time
import os
import logging
from openpyxl import load_workbook
import atlassian
import json
import requests
from numpy import nan
from typing import Dict, List, Any
from my_jira_app.utils import (
    MyLogger,
    Loader,
    search_import_file,
    # Can't use defaultdict due to pyqt conflicting with it
    # TypeError: multiple bases have instance lay-out conflict
    # MyDefaultDict,
    path_leaf,
)  # noqa: E501


class CreateTCS(MyLogger, Loader, QtCore.QRunnable):
    """
    This script creates new test cases on JIRA based on the mapping file.
    If there is no JIRA key attached to the test case and the process justifies
    the creation of a test case then a ticket is created.
    It also produces a .json file holding the grabbed data so that the sync.py
    script can later synchronize ARIS and JIRA without creating a new ticket
    Additionally, a consolidated excel + csv is also created to manually upload
    on JIRA in case the direct creation feature is turned off
    """

    def __init__(self, progressbar=None, statusBar=None) -> None:
        QtCore.QRunnable.__init__(self)
        self.progressbar = progressbar
        self.statusbar = statusBar
        self.parent = os.path.dirname(os.getcwd())
        self.yml_name = "generated_config_aris.yml"
        self.yml = self.load_yml(self.yml_name)
        # Namespace for the logger
        self.script_name = path_leaf(__file__)
        self.test_path = self.yml["test_path"]
        self.download_folder = self.yml["download_folder"]
        self.JIRA_project = self.yml["JIRA_project"]
        self.consolidated_file_name = self.yml["consolidated_file_name"]
        self.json_dump_file = self.yml["json_dump_file"]
        self.log_name = self.yml["log_name"]
        self.mapping_file_name = self.yml["mapping_file_name"]
        self.target = self.yml["target"]
        self.master = self.yml["master"]
        self.apps = self.yml["apps"]
        self.delete_all_xlsx = self.yml["delete_all_xlsx_files_from_temp_and_download"]
        self.delete_json_dump_file = self.yml["delete_json_dump_file"]
        self.clear_everything_create = self.yml["would_like_to_clear_everything_create"]

        super().__init__(log_name=self.log_name, name=self.script_name)

    def handle_QMessageBox_upstream_requests(self):
        if self.clear_everything_create:
            self.run_delete_all_xlsx()
            self.run_delete_json_dump_file()
        if self.delete_all_xlsx:
            self.run_delete_all_xlsx()
        if self.delete_json_dump_file:
            self.run_delete_json_dump_file()

    def run_delete_json_dump_file(self):
        dump_dir = os.path.join(self.parent, self.apps, self.json_dump_file)
        if os.path.isfile(dump_dir):
            os.remove(dump_dir)
            logging.getLogger(self.script_name).info(
                f"{self.json_dump_file} was deleted")
        else:
            logging.getLogger(self.script_name).info(
                f"tried to delete {self.json_dump_file} but could not be found")

    def run_delete_all_xlsx(self):
        xlsx_dir = os.path.join(self.parent, self.master, self.download_folder)
        if os.path.isdir(xlsx_dir):
            xlsx = search_import_file(".xlsx", where=xlsx_dir)
            list(map(lambda x: os.remove(x), xlsx))
            if len(xlsx) == 0:
                logging.getLogger(self.script_name).info(
                    "tried to delete the xlsx files but they could not be found")
            else:
                logging.getLogger(self.script_name).info(
                    f"{len(xlsx)} xlsx files were deleted")

    @staticmethod
    def mash_data(process: pd.DataFrame) -> dict:
        """
        Go through each line, store the group, obj name, obj type identifier and index of the bloc
        Build a temp json structure to hold relevant information of the bloc inside the try block

        "Is_predecessor_of","creates", "Activates", "Leads_to" indicate the link to the next item in the ARIS process
        whereas "Is_created_by","Is_activated_by", "follows", "evaluates", "Is_assigned_to" indicate the link to the previous item in the ARIS process
        """
        final_json: Dict[str, dict] = dict()
        prev_index = 0

        roles = {}
        try:
            roles = process.iloc[:, 1][process.iloc[:, 2] == "Role"].to_dict()
            roles = {v: k for k, v in roles.items()}
        except IndexError:
            pass

        for key, value in roles.items():
            k = 0
            attribute = process.iloc[value, 0]
            while attribute != "Remark/Example":
                k += 1
                if value + k == process.shape[0]:
                    break
                attribute = process.iloc[value + k, 0]

            owner = dict()
            if value + k == process.shape[0]:
                owner["owner"] = ""
            else:
                owner["owner"] = process.iloc[value + k, 1]
            owner["id"] = value
            roles[key] = owner

        for index, row in process.iterrows():

            trash_json = dict()

            if row[0] == "Group":
                _group = process.iloc[index, 1]
                _obj_name = process.iloc[index - 4, 0]
                _obj_type = process.iloc[index - 1, 1]
                _Identifier = process.iloc[index + 1, 1]
                prev_index = index
                if final_json.get(_Identifier):
                    _Identifier += "bis"

            try:
                trash_json["Group"] = _group
            except NameError:
                pass
            except UnboundLocalError:
                pass
            except KeyError:
                pass
            try:
                trash_json["Index"] = (index + 1, prev_index)
            except NameError:
                pass
            except UnboundLocalError:
                pass
            except KeyError:
                pass
            try:
                trash_json["Object Type"] = _obj_type
            except NameError:
                pass
            except UnboundLocalError:
                pass
            except KeyError:
                pass

            test_step = process[prev_index: index + 1]
            if test_step.shape[1] < 3:
                continue

            third_column = test_step.iloc[:, 2]
            first_column = test_step.iloc[:, 0]

            activates = first_column[first_column ==
                                     "activates"].dropna(how="all")
            if activates.empty is False:
                activate_index = activates.index.to_list()
                try:
                    trash_json["Activates"] = "§".join(
                        test_step.loc[activate_index, 1].tolist()
                    )
                except NameError:
                    pass
                except UnboundLocalError:
                    pass
                except KeyError:
                    pass
            is_predecessor_of = first_column[
                first_column == "is predecessor of"
            ].dropna(how="all")
            if is_predecessor_of.empty is False:
                is_predecessor_of_index = is_predecessor_of.index.to_list()
                try:
                    trash_json["Is_predecessor_of"] = "§".join(
                        test_step.loc[is_predecessor_of_index, 1].tolist()
                    )
                except NameError:
                    pass
                except UnboundLocalError:
                    pass
                except KeyError:
                    pass
            follows = first_column[first_column ==
                                   "follows"].dropna(how="all")
            if follows.empty is False:
                follows_index = follows.index.to_list()
                try:
                    trash_json["follows"] = "§".join(
                        test_step.loc[follows_index, 1].tolist()
                    )
                except NameError:
                    pass
                except UnboundLocalError:
                    pass
                except KeyError:
                    pass
            creates = first_column[first_column ==
                                   "creates"].dropna(how="all")
            if creates.empty is False:
                creates_index = creates.index.to_list()
                try:
                    trash_json["creates"] = "§".join(
                        test_step.loc[creates_index, 1].tolist()
                    )
                except NameError:
                    pass
                except UnboundLocalError:
                    pass
                except KeyError:
                    pass
            evaluates = first_column[first_column ==
                                     "evaluates"].dropna(how="all")
            if evaluates.empty is False:
                evaluates_index = evaluates.index.to_list()
                try:
                    trash_json["evaluates"] = "§".join(
                        test_step.loc[evaluates_index, 1].tolist()
                    )
                except NameError:
                    pass
                except UnboundLocalError:
                    pass
                except KeyError:
                    pass
            leads_to = first_column[first_column ==
                                    "leads to"].dropna(how="all")
            if leads_to.empty is False:
                leads_to_index = leads_to.index.to_list()
                try:
                    trash_json["Leads_to"] = "§".join(
                        test_step.loc[leads_to_index, 1].tolist()
                    )
                except NameError:
                    pass
                except UnboundLocalError:
                    pass
                except KeyError:
                    pass

            Is_assigned_to = first_column[first_column == "is assigned to"].dropna(
                how="all"
            )
            if Is_assigned_to.empty is False:
                Is_assigned_to_index = Is_assigned_to.index.to_list()
                try:
                    trash_json["Is_assigned_to"] = "§".join(
                        test_step.loc[Is_assigned_to_index, 1].tolist()
                    )
                except NameError:
                    pass
                except UnboundLocalError:
                    pass
                except KeyError:
                    pass
            is_activated_by = first_column[
                first_column == "is activated by"
            ].dropna(how="all")
            if is_activated_by.empty is False:
                is_activated_by_index = is_activated_by.index.to_list()
                try:
                    trash_json["Is_activated_by"] = "§".join(
                        test_step.loc[is_activated_by_index, 1].tolist()
                    )
                except NameError:
                    pass
                except UnboundLocalError:
                    pass
                except KeyError:
                    pass

            is_created_by = first_column[first_column == "is created by"].dropna(
                how="all"
            )
            if is_created_by.empty is False:
                is_created_by_index = is_created_by.index.to_list()
                try:
                    trash_json["Is_created_by"] = "§".join(
                        test_step.loc[is_created_by_index, 1].tolist()
                    )
                except NameError:
                    pass
                except UnboundLocalError:
                    pass
                except KeyError:
                    pass
            """
            Look for Role data

            """
            # drop row if all values are empty
            role = third_column[third_column == "Role"].dropna(how="all")
            if role.empty is False:
                role_index = role.index.to_list()[0]
                current_role = test_step.loc[role_index, 1]
                query_role_owner = roles.get(current_role)["owner"]
                if query_role_owner is not None:
                    try:
                        trash_json["Role"] = current_role + \
                            " - " + query_role_owner
                    except NameError:
                        pass
                    except UnboundLocalError:
                        pass
                    except KeyError:
                        pass
                else:
                    try:
                        trash_json["Role"] = current_role
                    except NameError:
                        pass
                    except UnboundLocalError:
                        pass
                    except KeyError:
                        pass
            """
            Look for Application system data

            """
            apps = first_column[first_column ==
                                "is supported by"].dropna(how="all")
            if apps.empty is False:
                app_index = apps.index.to_list()[0]
                current_app = test_step.loc[app_index, 1]
                try:
                    trash_json["Application_system"] = current_app
                except NameError:
                    pass
                except UnboundLocalError:
                    pass
                except KeyError:
                    pass
            try:
                if _obj_name == "4-eye principle":
                    continue
            except UnboundLocalError:
                pass
            try:
                trash_json["Object Name"] = _obj_name
                final_json[_Identifier] = trash_json
            except NameError:
                pass
            except UnboundLocalError:
                pass
            except KeyError:
                pass

        return final_json

    @staticmethod
    def convert_to_df(final_json: dict) -> pd.DataFrame:
        df = pd.DataFrame.from_dict(final_json, orient="index")
        next_columns = [
            column
            for column in ["Is_predecessor_of", "creates", "Activates", "Leads_to"]
            if column in df.columns
        ]
        previous_columns = [
            column
            for column in [
                "Is_created_by",
                "Is_activated_by",
                "follows",
                "evaluates",
                "Is_assigned_to",
            ]
            if column in df.columns
        ]
        df["next_value"] = df[next_columns].apply(
            lambda x: "§".join(x.dropna().astype(str)), axis=1
        )
        df["previous_value"] = df[previous_columns].apply(
            lambda x: "§".join(x.dropna().astype(str)), axis=1
        )

        return df

    @staticmethod
    def complement_missing_data(df: pd.DataFrame) -> pd.DataFrame:
        """
        Complement missing prev and next value data thanks to the information contained in the Rule entities
        """
        rules = df[df["Object Type"] == "Rule"]
        previous = rules.loc[:, ["previous_value"]]

        # ignore empty prev values
        non_empty_rows = previous.apply(lambda x: len(x.values[0]) > 0, axis=1)
        previous = previous[non_empty_rows]
        add = rules.loc[:, ["next_value"]]
        result = pd.concat([previous, add], axis=1,
                           sort=False).reindex(previous.index)

        # ignore empty next values
        non_empty_rows = result.apply(lambda x: len(x.values[1]) != 0, axis=1)
        result = result[non_empty_rows]

        for i, bubble_series in enumerate(result.iterrows()):
            rule_id, row = bubble_series
            prev_value = row["previous_value"]
            next_value = row["next_value"]
            found_prev_value = df[df["Object Name"] == prev_value]
            found_next_value = df[df["Object Name"] == next_value]

            # Insert Rule info into the event or function
            # Grab next value object add prev rule into the prev field
            if len(found_next_value.index) == 0 and isinstance(next_value, str) is True:
                list_of_next_values = next_value.split("§")
                for _next in list_of_next_values:
                    found = df[df["Object Name"] == _next]
                    if len(found.index) != 0:
                        obj_id = found.index.to_list()[0]
                        df.at[obj_id, "previous_value"] = prev_value
                    if len(found_prev_value.index) != 0:
                        obj_id = found_prev_value.index.to_list()[0]
                        df.at[obj_id, "next_value"] = _next

            # insert approximation from a rule into an event when event data is missing
            # this is an edge case
            elif isinstance(next_value, str) is True:
                obj_id = found_next_value.index.to_list()[0]
                if len(df.loc[obj_id, "previous_value"]) == 0:
                    df.loc[obj_id, "previous_value"] = "%"

            # Insert Rule info into the event or function
            # Grab prev value object add next rule into the next field
            if len(found_prev_value.index) == 0 and isinstance(prev_value, str) is True:
                list_of_previous_values = prev_value.split("§")
                for _prev in list_of_previous_values:
                    found = df[df["Object Name"] == _prev]
                    if len(found.index) != 0:
                        obj_id = found.index.to_list()[0]
                        df.at[obj_id, "next_value"] = next_value
                    if len(found_next_value.index) != 0:
                        _obj_id = found_next_value.index.to_list()[0]
                        df.at[_obj_id, "previous_value"] = _prev
        return df

    def grab_first_item(self, df: pd.DataFrame) -> List[str]:
        """
        returns ID of the first item in the process

        """
        _df = df.loc[:, ["Object Type", "previous_value",
                         "Object Name", "next_value"]]
        previous_values = _df[["previous_value", "Object Name", "next_value"]][
            _df["Object Type"].isin(["Function", "Event"])
        ]
        subs = df.loc[:, "Object Name"]
        first_items = list()
        for sub in subs:
            if "MCO" in sub:
                first_items.append(sub)

        previous_values = previous_values[previous_values["previous_value"] == ""]
        previous_values = previous_values[
            previous_values["next_value"].apply(lambda x: len(x)) > 0
        ]
        previous_values = previous_values[
            previous_values["Object Name"].isin(first_items)
        ]
        ordering = previous_values.index.tolist()
        assert isinstance(ordering, list)
        if len(ordering) > 1:
            logging.getLogger(self.script_name).warning(
                "Warning, multiple initial values found"
            )
        return ordering

    @staticmethod
    def apply_ordering(df: pd.DataFrame, ordering: list) -> pd.DataFrame:

        i = 0
        added: List[str] = list()

        for counter in range(df.shape[0]):
            if i < len(ordering):
                next_val = df.loc[ordering[i], "next_value"]
                """
                1) If the next value is empty the linear logic is broken

                This next loop will iterate over all ordering list elements and look for them in the previous_value column
                In the example below line 2 is broken as it leads nowhere.
                What this next loop will do is look at elements in the ordering list so far [A, B]
                and add the rows that have a previous value that matches any of the ordering list elements

                ex: 3x3 matrix with 3 columns: previous_value, Object Name and next_value

                previous_value	Object Name	next_value

                X	A	B
                Y	B
                A	C


                2) If the next value was already added to the ordering list the linear logic is broken to avoid circular relationships

                """

                if len(next_val) == 0 or next_val in added:
                    sanity = 1
                    idx1 = pd.Index(ordering)
                    idx2 = df.index
                    diff = idx2.difference(idx1)
                    temp = df[["Object Name", "next_value", "previous_value"]][
                        df["Object Type"].isin(["Function", "Event"])
                    ]
                    temp = temp[temp["next_value"] != ""]
                    temp = temp[temp["previous_value"] != ""]
                    next_val = temp[temp.index.isin(diff)]
                    next_val = df.loc[ordering[i - sanity], "Object Name"]
                    copy = len(added)
                    while len(added) == copy or i - sanity > 0:

                        next_val = df.loc[ordering[i - sanity], "Object Name"]
                        sanity += 1
                        list_of_next_values = next_val.split("§")
                        for next_valu in list_of_next_values:
                            found = df[df["previous_value"] == next_valu]
                            if len(found.index) != 0:
                                obj_list = found.index.to_list()
                                for obj_id in obj_list:
                                    if obj_id not in ordering:
                                        ordering.append(obj_id)
                                        added.append(next_val)
                                        i += 1
                        if i - sanity < 0:
                            break

                else:
                    """
                    if the next value exists proceed as expected: linear progression

                    """
                    list_of_next_values = next_val.split("§")
                    for next_valu in list_of_next_values:
                        found = df[df["Object Name"] == next_valu]
                        if len(found.index) != 0:
                            obj_list = found.index.to_list()
                            for obj_id in obj_list:
                                if obj_id not in ordering:
                                    ordering.append(obj_id)
                                    added.append(next_val)
                                    i += 1

        """
        Complement missing index values in case the process above dismissed entities

        """
        original_index_values = df.index.to_list()
        for original_index_value in original_index_values:
            if original_index_value not in ordering:
                ordering.append(original_index_value)
        if "Application_system" in df.columns:
            df["Object Name"] = df[["Object Name", "Application_system"]].apply(
                lambda x: " on ".join(x.dropna().astype(str)), axis=1
            )
            to_keep = df["Application_system"].dropna().index.to_list()
            df = df.loc[to_keep]
            result = df.loc[
                :,
                [
                    "Object Type",
                    "Role",
                    "previous_value",
                    "Object Name",
                    "next_value",
                    "Application_system",
                ],
            ]
            result = result.reindex(ordering)
            result = result[result["Object Type"].isin(["Function"])]
            result = result.loc[
                :,
                [
                    "Role",
                    "previous_value",
                    "Object Name",
                    "next_value",
                    "Application_system",
                ],
            ]
        else:
            result = pd.DataFrame()

        return result

    @staticmethod
    def remove_duplicates(x: Dict[str, Dict[Any, Any]]) -> List[str]:
        return list(dict.fromkeys(x))

    @staticmethod
    def grab_u_number(business_rep: str) -> str:
        if business_rep == "Schneider, Christian":
            return "u54119"
        elif business_rep == "Koch, Heiko":
            return "u54071"
        elif business_rep == "Moritz, Christian":
            return "u54090"
        elif business_rep == "Schmitz, Christian":
            return "u54114"
        elif business_rep == "Mouton, Jeff":
            return "u46435"
        elif business_rep == "August, Amela":
            return "u54063"
        elif business_rep == "Heimhard, Ulrich":
            return "u54058"
        elif business_rep == "Baeumer, Torsten":
            return "u54006"
        elif business_rep == "Fabry, Gerard":
            return "u53116"
        elif business_rep == "Broermann, Lars":
            return "u41261"
        elif business_rep == " Dennis Schwarz":
            return "u41514"
        elif business_rep == "Gaelle Migani":
            return "u43896"
        elif business_rep == "Migani, Gaelle":
            return "u43896"
        elif business_rep == "Demartini, Laurent":
            return "u43392"
        elif business_rep == "Ziegler, Martin":
            return "u24712"
        elif business_rep == "Battisti, Nicolas":
            return "u34904"
        elif business_rep == "Rea, Marie-Noelle":
            return "u55315"
        elif business_rep == "Raoux, Marie":
            return "u55314"
        elif business_rep == "Schneider, Laurent":
            return "u55317"
        elif business_rep == "Florio, Nicolas":
            return "u37000"

        return f"No mapping id for user: '{business_rep}'"

    def convert_to_xray_format(
        self,
        df: pd.DataFrame,
        file: str,
        test_case_num: int,
        path_to_mapping: str,
    ) -> pd.DataFrame:

        applications = ""
        mapping = pd.read_excel(path_to_mapping, sheet_name="mapping")
        xray = pd.DataFrame()
        b_rep_df = pd.read_excel(file)
        business_rep = b_rep_df.iloc[:, 1][
            b_rep_df.iloc[:, 0] == "Process Owner_BJB"
        ].to_list()

        if len(business_rep) != 0:
            business_rep = business_rep[0]
            u_num_business_rep = self.grab_u_number(business_rep)
            if u_num_business_rep is None:
                u_num_business_rep = "u46022"
        else:
            u_num_business_rep = "u46022"

        if "Application_system" in df.columns:
            xray = df.loc[:, "Application_system"]
            applications = xray.dropna().to_list()
            assert isinstance(applications, list)
            applications = self.remove_duplicates(applications)
            x_ray_columns = [
                "TCID",
                "Status",
                "Assignee",
                "Reporter",
                "Summary",
                "Description",
                "Responsible",
                "Affected Application/s",
                "Priority",
                "Test Step",
                "Expected result",
                "Location",
            ]
            for i in range(len(applications)):
                label = "Label" + str(i+1)
                x_ray_columns.append(label)

            applications_str = ", ".join(applications)

            first_index = df.index.to_list()[0]
            xray = pd.DataFrame(
                columns=x_ray_columns
            )
            # xray['Expected result'] = df.loc[:,"next_value"]
            xray.loc[:, "Test Step"] = df.loc[:, "Object Name"]
            xray.loc[:, "TCID"] = test_case_num
            xray.loc[first_index, "Status"] = "DRAFT"  # "IN USE"
            file = path_leaf(file)
            filename, file_extension = os.path.splitext(file)
            #            timerz = time.strftime("_%Y%m%d")
            tree_element = mapping[mapping["Process_filename"] == file][
                "Tree"
            ].to_list()
            if len(tree_element) > 0:
                tree_element = tree_element[0]
                xray.loc[first_index, "Location"] = (
                    self.test_path
                    + "/"
                    + tree_element
                    + "/"
                    + filename.replace("/", "-")
                )
            xray.loc[first_index, "Assignee"] = u_num_business_rep
            xray.loc[first_index, "Reporter"] = u_num_business_rep
            xray.loc[first_index, "Summary"] = filename
            # xray.loc[first_index,'Description'] = file
            xray.loc[first_index, "Responsible"] = u_num_business_rep
            xray.loc[first_index, "Affected Application/s"] = applications_str
            xray.loc[first_index, "Priority"] = "Major"
            for i, application in enumerate(applications):
                label = "Label" + str(i+1)
                xray.loc[first_index, label] = application

        return xray

    @staticmethod
    def generate_excel(df: pd.DataFrame, file_name: str) -> None:
        timerz = time.strftime("%Y%m%d_%H%M%S")
        filename = file_name + "_" + timerz + ".xlsx"
        # compress all dataframes into a single dataframe for xray import
        writer = pd.ExcelWriter(
            filename, engine="xlsxwriter", options={"remove_timezone": True}
        )
        df.to_excel(writer, sheet_name="Data", header=True, index=False)
        writer.save()

    def generate_csv(self, df: pd.DataFrame, file_name: str) -> None:
        timerz = time.strftime("_%Y%m%d_%H%M%S")
        # compress all dataframes into a single dataframe for xray import
        name = os.path.join(os.getcwd(), file_name + timerz + ".csv")
        df.to_csv(name, sep=";", header=True, index=False)

    def add_url(
        self, df: pd.DataFrame, path_to_mapping: str, cur_file: str
    ) -> pd.DataFrame:
        mapping = pd.read_excel(path_to_mapping, sheet_name="mapping")
        url = mapping["Unnamed: 0"][mapping["Abs_path"] == cur_file]

        if df.empty is False:
            url = url.to_list()
            if len(url) > 0:
                url = url[0]
                first_index = df.index.to_list()[0]
                df.loc[first_index, "Description"] = url.replace(" ", "%20")
            else:
                first_index = df.index.to_list()[0]
                df.loc[first_index, "Description"] = "could not retrieve url"
                logging.getLogger(self.script_name).info(
                    f"{path_leaf(cur_file)} abs_path could not be found on the mapping file, skipping"
                )

        return df

    def compress_into_one_file(self) -> pd.DataFrame:
        path_to_files = search_import_file(
            where=os.path.join(self.parent, self.master, self.download_folder),
            extension=".xlsx",
        )
        consolidated = pd.DataFrame()
        for path_to_file in path_to_files:
            dummy = pd.read_excel(path_to_file, sheet_name="Data")
            consolidated = pd.concat([consolidated, dummy], sort=False)
        return consolidated

    @staticmethod
    def add_jira_key(jira_key: str, path_to_mapping: str, cur_file: str) -> None:
        mapping = pd.read_excel(path_to_mapping, sheet_name="mapping")
        row = mapping[mapping["Abs_path"] == cur_file].index.to_list()
        if len(row) > 0:
            row = row[0]
            wb = load_workbook(path_to_mapping)
            sheets = wb.sheetnames
            reports = wb[sheets[0]]
            reports.cell(row=int(row) + 2, column=3).value = jira_key
            wb.save(path_to_mapping)

    @staticmethod
    def build_test_data(df: pd.DataFrame) -> dict:
        if df.empty is False:
            test_list = df["Test Step"].to_list()
            list_of_dict = list()
            for i, test in enumerate(test_list):
                list_of_dict.append({"index": i + 1, "step": test})
            return {"steps": list_of_dict}
        return {"steps": "empty"}

    def create_jira_ticket_payload(self, df: pd.DataFrame, file: str) -> dict:
        wrapper_of_wrapper = self.build_test_data(df)

        if df.empty is False:
            first_index = df.index.to_list()[0]
            df.loc[first_index, "Summary"]
            payload = {
                "project": {"key": self.JIRA_project},
                "summary": df.loc[first_index, "Summary"],
                "assignee": {"name": df.loc[first_index, "Assignee"]},
                "reporter": {"name": df.loc[first_index, "Reporter"]},
                "priority": {"name": df.loc[first_index, "Priority"]},
                # Test Repository Path: customfield_17291
                "customfield_17291": df.loc[first_index, "Location"],
                "description": df.loc[first_index, "Description"],
                # "customfield_15380" "responsible"
                "customfield_15380": {"name": df.loc[first_index, "Responsible"]},
                # "Affected Application/s" "customfield_16181"
                "customfield_16181": [df.loc[first_index, "Affected Application/s"]],
                "issuetype": {"name": "Test"},
                # Manual test steps: customfield_17284
                "customfield_17284": wrapper_of_wrapper,
            }
            return payload
        else:
            logging.getLogger(self.script_name).error(f"File is empty: {file}")
            return {}

    def create_jira_ticket(self, payload: dict) -> str:
        loader = Loader(self.yml_name)
        jira = atlassian.Jira(
            url=self.target, username=loader.cfg.user, password=loader.cfg.password,
        )
        response = jira.issue_create(fields=payload)
        if "key" in response.keys():
            key = response["key"]
        else:
            error_occured = f"""
            jira ticket not created. JIRA target server
            ({self.target}) could be down or this could be due to a
            permission issue on JIRA. Check the logs to debug.
            Type "skip" to cancel the creation or press ENTER to retry
            """
            # POP UP HERE
            self.statusbar.showMessage(
                "Program paused, please refer to the console")
            logging.getLogger(self.script_name).error(error_occured)
            if "errors" in response.keys():
                logging.getLogger(self.script_name).error(
                    f"{response['errors']}")
            skip = input(error_occured)
            if skip != "skip":
                response = jira.issue_create(fields=payload)
                key = response["key"]
            else:
                return ""
        #        jira.issue_transition(key, "In Use")
        assert isinstance(key, str)
        logging.getLogger(self.script_name).info(f"{key} created")
        return key

    def ticket_does_not_exist(self, path_to_mapping: str, cur_file: str) -> bool:
        mapping = pd.read_excel(path_to_mapping, sheet_name="mapping")
        key = mapping["JIRA_key"][mapping["Abs_path"] == cur_file]

        if key.empty or key.values[0] is nan:
            logging.getLogger(self.script_name).info(
                f"{path_leaf(cur_file)} not found, calling JIRA webservice for creation"
            )
            return True
        elif key.values[0] is not nan and isinstance(key.values[0], str):
            if len(key.values[0]) > 0:
                logging.getLogger(self.script_name).info(
                    f"{path_leaf(cur_file)} found on the mapping file, cancel creation"
                )
                return False
        return True

    @staticmethod
    def check_file_size(path_to_file: str) -> bool:
        size = os.path.getsize(path_to_file)
        if size < 100:
            return False
        return True

    def save_json(self, data: Dict[str, dict]) -> list:
        payloads: List[Dict[str, dict]] = self.load_json()
        already_in_json = False
        for key, value in data.items():
            key_data = key
        if len(payloads) != 0 and payloads != "[]":
            for i in payloads:
                for key, value in i.items():
                    assert isinstance(key, str)
                    if key == key_data:
                        already_in_json = True
        assert isinstance(data, dict)
        if already_in_json is False:
            payloads.append(data)
        logging.getLogger(self.script_name).info(
            f"{len(payloads)} processes were saved on {self.json_dump_file}"
        )
        with open(self.json_dump_file, "w") as f:
            json.dump(payloads, f)
        return payloads

    def load_json(self) -> List[Dict[str, dict]]:
        if os.path.isfile(self.json_dump_file) is False:
            with open(self.json_dump_file, "w") as f:
                data = "[]"
                f.write(data)
                logging.getLogger(self.script_name).info(
                    f"{self.json_dump_file} does not exist creating file"
                )
                f.close()
                return list()

        with open(self.json_dump_file, "r") as f:
            data = json.load(f)
            logging.getLogger(self.script_name).info(
                f"{len(data)} processes were loaded from {self.json_dump_file}"
            )
            assert isinstance(data, list)
            return data

    def update_label_jira_ticket(self, df, jira_key):
        label_df = df.loc[:, df.columns.str.contains('Label')]
        first_index = label_df.index.to_list()[0]
        loader = Loader(self.yml_name)
        headers = {"Authorization": "Basic %s" % loader.cfg.u,
                   "Content-Type": "application/json"}
        for i in range(label_df.shape[1]):
            label = "Label"+str(i+1)
            label_value = str(df.loc[first_index, label]).replace(" ", "")
            response = requests.put(self.target + 'rest/api/2/issue/' + str(jira_key), data=json.dumps(
                {"update": {"labels": [{"add": label_value}]}}), headers=headers, verify=False)

            if len(response.text) > 0:
                logging.getLogger(self.script_name).error(
                    f"{response} could not add label to key {jira_key}"
                )
                logging.getLogger(self.script_name).error(
                    f"{response.text}"
                )

    def init_progress_counter(self, path_to_mapping):
        mapping = pd.read_excel(path_to_mapping, sheet_name="mapping")
        self.total_tickets = mapping.shape[0]
        # substract already created tickets
        already_created = len(mapping["JIRA_key"]
                              [mapping["JIRA_key"].isnull() == False])
        self.total_tickets -= already_created
        logging.getLogger(self.script_name).info(
            f"{self.total_tickets} will potentially be created if the conditions apply, {already_created} are already created so they will be skipped"
        )

    def process_files(self) -> list:
        path_to_mapping = os.path.join(
            self.parent, self.master, self.mapping_file_name)
        files = search_import_file(
            where=os.path.join(self.parent, self.master, self.download_folder),
            extension=".xls",
        )
        keep_track_test_case_num = 1
        data = dict()
        self.init_progress_counter(path_to_mapping)

        for file in files:

            logging.getLogger(self.script_name).info(
                f"Processing started for: {file}")
            path_to_file = os.path.join(
                self.parent, self.master, self.download_folder, file
            )
            if self.check_file_size(path_to_file):
                process = pd.read_excel(
                    path_to_file,
                    encoding="iso-8859-1",
                    header=None
                )

                final_json = self.mash_data(process)
                df = self.convert_to_df(final_json)
                if (
                    self.ticket_does_not_exist(path_to_mapping, file)
                    and df.empty is False
                ):
                    df = self.complement_missing_data(df)
                    ordering = self.grab_first_item(df)
                    df = self.apply_ordering(df, ordering)
                    df = self.convert_to_xray_format(
                        df,
                        file,
                        keep_track_test_case_num,
                        path_to_mapping,
                    )
                    df = self.add_url(df, path_to_mapping, file)
                    payload = self.create_jira_ticket_payload(df, file)
                    if (
                        payload is not None
                        and "description" in payload.keys()
                    ):
                        _id = payload["description"]
                        data[_id] = payload
                        payloads = self.save_json(data)
                        jira_key = self.create_jira_ticket(payload)
                        self.add_jira_key(jira_key, path_to_mapping, file)
                        if len(jira_key) > 0:
                            self.update_label_jira_ticket(df, jira_key)
                        self.generate_excel(df, file)
                        keep_track_test_case_num += 1
                else:
                    logging.getLogger(self.script_name).info(
                        f"Webservice not called as there is nothing to create. The process on ARIS does not justify the test case creation"
                    )
            else:
                logging.getLogger(self.script_name).info(
                    f"{path_leaf(file)} is too small, skipping"
                )
            if self.total_tickets != 0:
                progress = (keep_track_test_case_num-1)*100/self.total_tickets
                self.update_progress_bar(progress, keep_track_test_case_num)

        return payloads

    def update_progress_bar(self, currentPercentage, num=1):
        if num-1 != 0:
            logging.getLogger(self.script_name).info(
                f"{num-1} test cases created"
            )
        if self.statusbar is not None:
            self.statusbar.showMessage(
                f"Creating tickets: {self.total_tickets-num} TCs left to process")
            QtCore.QMetaObject.invokeMethod(self.progressbar, "setValue",
                                            QtCore.Qt.QueuedConnection,
                                            QtCore.Q_ARG(
                                                int, currentPercentage))

    def run(self):
        self.handle_QMessageBox_upstream_requests()

        input(
            f"""
            - Move created tickets to archived on JIRA
            - Make sure the {self.mapping_file_name} file is closed
            - The file {self.mapping_file_name} will be populated with newly created {self.JIRA_project} tickets
            """
        )

        self.process_files()
        consolidated = self.compress_into_one_file()
        os.chdir(os.path.join(self.parent, self.master, "consolidated"))
        self.generate_csv(
            consolidated, self.consolidated_file_name)
        self.generate_excel(
            consolidated, self.consolidated_file_name
        )
        os.chdir(os.path.join(self.parent, self.apps))
        self.update_progress_bar(100)
        self.statusbar.showMessage("Done")


def main():
    creator_instance = CreateTCS()
    creator_instance.run()


if __name__ == "__main__":
    main()
