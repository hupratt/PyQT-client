import pandas as pd
import time, os, logging
from utils import MyDefaultDict
from openpyxl import load_workbook
import atlassian, json
from utils import Loader
from numpy import nan

def mash_data(process: pd.DataFrame) -> dict:
    """ 
    Go through each line, store the group, obj name, obj type identifier and index of the bloc
    Build a temp json structure to hold relevant information of the bloc inside the try block
    
    "Is_predecessor_of","creates", "Activates", "Leads_to" indicate the link to the next item in the ARIS process
    whereas "Is_created_by","Is_activated_by", "follows", "evaluates", "Is_assigned_to" indicate the link to the previous item in the ARIS process
    
    """
    final_json = dict()
    prev_index = 0
    
    roles = process[0][process[1] == "Role"].to_dict()
    roles = {v: k for k, v in roles.items()}
    
    for key, value in roles.items():
        k = 0
        attribute = process.iloc[value, 0]
        while attribute != "Remark/Example":
            k += 1
            attribute = process.iloc[value + k, 0]
        owner = dict()
        owner['owner'] = process.iloc[value + k, 1]
        owner['id'] = value
        roles[key] = owner
        
    for index, row in process.iterrows():
        
        trash_json = MyDefaultDict()
    
        if row[0] == "Group":
            _group = process.loc[index,1]
            _obj_name = process.loc[index-1,0]
            _obj_type = process.loc[index-1,1]
            _Identifier = process.loc[index+1,1]
            prev_index = index
            if final_json.get(_Identifier):
                _Identifier += "bis"
        
        try:
            trash_json["Group"] = _group
            trash_json["Index"] = (index + 1, prev_index)
            trash_json["Object Type"] = _obj_type
            
            test_step = process[prev_index: index+1]
            third_column = test_step[2]
            first_column = test_step[0]
    
            activates = first_column[first_column == "activates"].dropna(how='all')
            if activates.empty is False:
                activate_index = activates.index.to_list()
                trash_json["Activates"] = "§".join(test_step.loc[activate_index, 1].tolist()) 
            
            is_predecessor_of = first_column[first_column == "is predecessor of"].dropna(how='all')
            if is_predecessor_of.empty is False:
                is_predecessor_of_index = is_predecessor_of.index.to_list()
                trash_json["Is_predecessor_of"] = "§".join(test_step.loc[is_predecessor_of_index, 1].tolist())
                
            follows = first_column[first_column == "follows"].dropna(how='all')
            if follows.empty is False:
                follows_index = follows.index.to_list()
                trash_json["follows"] = "§".join(test_step.loc[follows_index, 1].tolist())
                
            creates = first_column[first_column == "creates"].dropna(how='all')
            if creates.empty is False:
                creates_index = creates.index.to_list()
                trash_json["creates"] = "§".join(test_step.loc[creates_index, 1].tolist())
                
            evaluates = first_column[first_column == "evaluates"].dropna(how='all')
            if evaluates.empty is False:
                evaluates_index = evaluates.index.to_list()
                trash_json["evaluates"] = "§".join(test_step.loc[evaluates_index, 1].tolist())
                
            leads_to = first_column[first_column == "leads to"].dropna(how='all')
            if leads_to.empty is False:
                leads_to_index = leads_to.index.to_list()  
                trash_json["Leads_to"] = "§".join(test_step.loc[leads_to_index, 1].tolist()) 

            Is_assigned_to = first_column[first_column == "is assigned to"].dropna(how='all')
            if Is_assigned_to.empty is False:
                Is_assigned_to_index = Is_assigned_to.index.to_list()
                trash_json["Is_assigned_to"] = "§".join(test_step.loc[Is_assigned_to_index, 1].tolist())
                
            is_activated_by = first_column[first_column == "is activated by"].dropna(how='all')
            if is_activated_by.empty is False:
                is_activated_by_index = is_activated_by.index.to_list()
                trash_json["Is_activated_by"] = "§".join(test_step.loc[is_activated_by_index, 1].tolist())
            
            is_created_by = first_column[first_column == "is created by"].dropna(how='all')
            if is_created_by.empty is False:
                is_created_by_index = is_created_by.index.to_list()
                trash_json["Is_created_by"] = "§".join(test_step.loc[is_created_by_index, 1].tolist())
    
            """
            Look for Role data
            
            """
            # drop row if all values are empty
            role = third_column[third_column == "Role"].dropna(how='all')
            if role.empty is False:
                role_index = role.index.to_list()[0]
                current_role = test_step.loc[role_index, 1]
                query_role_owner = roles.get(current_role)['owner']
                if query_role_owner is not None:
                    trash_json["Role"] = current_role  + " - " + query_role_owner
                else:
                    trash_json["Role"] = current_role 
                    
            """
            Look for Application system data
            
            """
            apps = first_column[first_column == "is supported by"].dropna(how='all')
            if apps.empty is False:
                app_index = apps.index.to_list()[0]
                current_app = test_step.loc[app_index, 1]
                trash_json["Application_system"] = current_app 
                    
            if _obj_name == "4-eye principle":
                continue
            
            trash_json["Object Name"] = _obj_name  
            final_json[_Identifier] = trash_json
            
        except NameError:
            pass
        
    return final_json
    

def convert_to_df(final_json: dict) -> pd.DataFrame:
    df = pd.DataFrame.from_dict(final_json,orient='index')
    next_columns = [column for column in ["Is_predecessor_of","creates", "Activates", "Leads_to"] if column in df.columns]
    previous_columns = [column for column in ["Is_created_by","Is_activated_by", "follows", "evaluates", "Is_assigned_to"] if column in df.columns]
    df['next_value'] = df[next_columns].apply(lambda x: '§'.join(x.dropna().astype(str)), axis=1)
    df['previous_value'] = df[previous_columns].apply(lambda x: '§'.join(x.dropna().astype(str)), axis=1)
    
    return df
  


def complement_missing_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    Complement missing prev and next value data thanks to the information contained in the Rule entities
    
    """
    rules = df[df["Object Type"]=="Rule"]
    previous = rules.loc[:,["previous_value"]]
    
    # ignore empty prev values
    non_empty_rows = previous.apply(lambda x: len(x.values[0]) > 0, axis=1)
    previous = previous[non_empty_rows]
    add = rules.loc[:,["next_value"]]
    result = pd.concat([previous, add], axis=1, sort=False).reindex(previous.index)
    
    # ignore empty next values
    non_empty_rows = result.apply(lambda x: len(x.values[1]) != 0, axis=1)
    result = result[non_empty_rows]
    
    for i, bubble_series in enumerate(result.iterrows()):
        rule_id, row = bubble_series   
        prev_value = row["previous_value"]
        next_value = row["next_value"]
        found_prev_value = df[df["Object Name"] == prev_value]
        found_next_value = df[df["Object Name"] == next_value]
        
        # Insert Rule info into the event or function
        # Grab next value object add prev rule into the prev field
        if len(found_next_value.index) == 0 and isinstance(next_value,str) is True: 
            list_of_next_values = next_value.split("§")
            for _next in list_of_next_values:
                found = df[df["Object Name"]==_next]
                if len(found.index) != 0:
                    obj_id = found.index.to_list()[0]
                    df.at[obj_id,"previous_value"] = prev_value
                if len(found_prev_value.index) != 0:
                    obj_id = found_prev_value.index.to_list()[0]
                    df.at[obj_id, "next_value"] = _next
        
        # insert approximation from a rule into an event when event data is missing
        # this is an edge case
        elif isinstance(next_value,str) is True: 
            obj_id = found_next_value.index.to_list()[0]
            if len(df.loc[obj_id,"previous_value"]) == 0:
                df.loc[obj_id,"previous_value"] = "%"
    
        # Insert Rule info into the event or function
        # Grab prev value object add next rule into the next field
        if len(found_prev_value.index) == 0 and isinstance(prev_value,str) is True: 
            list_of_previous_values = prev_value.split("§")
            for _prev in list_of_previous_values:
                found = df[df["Object Name"]==_prev]
                if len(found.index) != 0:
                    obj_id = found.index.to_list()[0]
                    df.at[obj_id, "next_value"] = next_value
                if len(found_next_value.index) != 0:
                    _obj_id = found_next_value.index.to_list()[0]
                    df.at[_obj_id,"previous_value"] = _prev
    return df


def grab_first_item(df: pd.DataFrame) -> list:
    """ 
    returns ID of the first item in the process 
    
    """
    _df = df.loc[:,["Object Type", "previous_value", "Object Name", "next_value"]]
    previous_values = _df[["previous_value","Object Name", "next_value"]][_df["Object Type"].isin(["Function","Event"])]
    subs = df.loc[:,"Object Name"]
    first_items = list()
    for sub in subs:
        if "MCO" in sub:
            first_items.append(sub)
            
    previous_values = previous_values[previous_values["previous_value"] == ""]
    previous_values = previous_values[previous_values["next_value"].apply(lambda x: len(x)) > 0]
    previous_values = previous_values[previous_values["Object Name"].isin(first_items)]
    ordering = previous_values.index.tolist()
    if len(ordering) > 1:
        logging.getLogger(__name__).error("Warning, multiple initial values found")
    return ordering
    
def apply_ordering(df: pd.DataFrame, ordering: list) -> pd.DataFrame:
    
    i = 0
    added = list()
    
    for counter in range(df.shape[0]):
        if i < len(ordering):
            next_val = df.loc[ordering[i],"next_value"]
            """
            1) If the next value is empty the linear logic is broken
    
            This next loop will iterate over all ordering list elements and look for them in the previous_value column
            In the example below line 2 is broken as it leads nowhere. 
            What this next loop will do is look at elements in the ordering list so far [A, B] 
            and add the rows that have a previous value that matches any of the ordering list elements
            
            ex: 3x3 matrix with 3 columns: previous_value, Object Name and next_value
            
            previous_value	Object Name	next_value
            		
            X	A	B
            Y	B	
            A	C	
            
            
            2) If the next value was already added to the ordering list the linear logic is broken to avoid circular relationships
            
            """
                
            if len(next_val) == 0 or next_val in added:
                sanity = 1
                idx1 = pd.Index(ordering)
                idx2 = df.index
                diff = idx2.difference(idx1)
                temp = df[["Object Name", "next_value", "previous_value"]][df["Object Type"].isin(["Function","Event"])]
                temp = temp[temp["next_value"] != ""]
                temp = temp[temp["previous_value"] != ""]
                next_val = temp[temp.index.isin(diff)]
                next_val = df.loc[ordering[i-sanity],"Object Name"]
                copy = len(added)
                while len(added) == copy or i-sanity > 0:
                    
                    next_val = df.loc[ordering[i-sanity],"Object Name"]
                    sanity += 1
                    list_of_next_values = next_val.split("§")
                    for next_valu in list_of_next_values:
                        found = df[df["previous_value"] == next_valu]
                        if len(found.index) != 0:
                            obj_list = found.index.to_list()
                            for obj_id in obj_list:
                                if obj_id not in ordering:
                                    ordering.append(obj_id)
                                    added.append(next_val)
                                    i += 1
                    if i-sanity < 0:
                        break   

            else:
                """
                if the next value exists proceed as expected: linear progression
                
                """
                list_of_next_values = next_val.split("§")
                for next_valu in list_of_next_values:
                    found = df[df["Object Name"] == next_valu]
                    if len(found.index) != 0:
                        obj_list = found.index.to_list()
                        for obj_id in obj_list:
                            if obj_id not in ordering:
                                ordering.append(obj_id)
                                added.append(next_val)
                                i += 1            
                                
    """
    Complement missing index values in case the process above dismissed entities
    
    """
    original_index_values = df.index.to_list()
    for original_index_value in original_index_values:
        if original_index_value not in ordering:
            ordering.append(original_index_value)
    if "Application_system" in df.columns:
        df['Object Name'] = df[['Object Name', "Application_system"]].apply(lambda x: ' on '.join(x.dropna().astype(str)), axis=1)
        to_keep = df["Application_system"].dropna().index.to_list()
        df = df.loc[to_keep]
        result = df.loc[:,["Object Type", "Role", "previous_value", "Object Name", "next_value", "Application_system"]]
        result = result.reindex(ordering)
        result = result[result["Object Type"].isin(["Function"])]
        result = result.loc[:,["Role", "previous_value", "Object Name", "next_value", "Application_system"]]
    else:  
        result = pd.DataFrame()
     
    return result

def remove_duplicates(x):
  return list(dict.fromkeys(x))

def convert_to_xray_format(df: pd.DataFrame, file: str, test_case_num: int, business_rep: pd.DataFrame) -> pd.DataFrame:
    applications = ""
    xray = pd.DataFrame()
    business_rep = business_rep[1][business_rep[0]=="     Responsible"].to_list()
    if len(business_rep) != 0:
        business_rep = business_rep[0]
    else:
        business_rep = 'u46022'
    
    if "Application_system" in df.columns:
        xray = df.loc[:,"Application_system"]
        applications = xray.dropna().to_list()
        applications = remove_duplicates(applications)
        applications = ", ".join(applications)
        
        first_index = df.index.to_list()[0]
        xray = pd.DataFrame(columns=["TCID","Status",'Assignee','Reporter','Summary','Description','Responsible','Affected Application/s','Priority','Test Step','Expected result', 'Location'])
        # xray['Expected result'] = df.loc[:,"next_value"]
        xray.loc[:,'Test Step'] = df.loc[:,"Object Name"]
        xray.loc[:,'TCID'] = test_case_num
        xray.loc[first_index,'Status'] = "IN USE"
        filename, file_extension = os.path.splitext(file)
        timerz = time.strftime("_%Y%m%d")
        xray.loc[first_index,'Location'] = "/Testing/milestone6" + timerz + "/" + filename.replace("/","-")
        xray.loc[first_index,'Assignee'] = 'u44714'
        xray.loc[first_index,'Reporter'] = 'u44714'
        xray.loc[first_index,'Summary'] = filename
        # xray.loc[first_index,'Description'] = file
        xray.loc[first_index,'Responsible'] = business_rep
        xray.loc[first_index,'Affected Application/s'] = applications
        xray.loc[first_index,'Priority'] = 'Major'
    
    return xray
    

def generate_excel(df: pd.DataFrame, file_name: str) -> None:
    timerz = time.strftime("%Y%m%d_%H%M%S")
    filename = file_name + "_" + timerz + '.xlsx'
    # compress all dataframes into a single dataframe for xray import
    writer = pd.ExcelWriter(filename, engine='xlsxwriter', options={'remove_timezone': True})
    df.to_excel(writer, sheet_name="Data", header=True, index=False)
    writer.save()

def generate_csv(df: pd.DataFrame, file_name: str) -> None:
    timerz = time.strftime("_%Y%m%d_%H%M%S")
    # compress all dataframes into a single dataframe for xray import   
    name = os.path.join(os.getcwd(), file_name + timerz + '.csv')
    df.to_csv(name, sep=';', header=True, index=False)
    
def add_url(df, path_to_mapping, cur_file):
    mapping = pd.read_excel(path_to_mapping, sheet_name="Reports")
    url = mapping["url #1"][mapping["Filename"] == cur_file]
    url = url.to_list()[0]
    if df.empty is False:
        first_index = df.index.to_list()[0]
        df.loc[first_index,'Description'] = url
    return df

def compress_into_one_file():
    path_to_files = search_import_file(".xlsx")
    consolidated = pd.DataFrame()
    for path_to_file in path_to_files:
        dummy = pd.read_excel(path_to_file, sheet_name="Data")
        consolidated = pd.concat([consolidated, dummy])
    return consolidated
        
def search_import_file(extension):
    _directory = [i for i in os.listdir(os.path.curdir)]
    path_to_files = list()
    curdir = os.path.abspath(os.path.curdir)
    for file in _directory: 
        if file.endswith(extension):
            path_to_files.append(os.path.join(curdir,file))
    return path_to_files

def add_jira_key(jira_key, path_to_mapping, cur_file):
    mapping = pd.read_excel(path_to_mapping, sheet_name="Reports")
    row = mapping[mapping["Filename"] == cur_file].index.to_list()[0]
    wb = load_workbook(path_to_mapping)
    sheets = wb.sheetnames
    reports = wb[sheets[0]]
    reports.cell(row = int(row)+2, column = 8).value = jira_key
    wb.save(path_to_mapping)

def build_test_data(df):
    if df.empty is False:
        test_list = df['Test Step'].to_list()
        list_of_dict = list()
        for i, test in enumerate(test_list):
            list_of_dict.append({"index":i+1, "step": test})
        return {"steps":list_of_dict}
    

def create_ticket_payload(df, file):
    
    wrapper_of_wrapper = build_test_data(df)
    
    if df.empty is False:
        first_index = df.index.to_list()[0]
        df.loc[first_index,'Summary']
        payload = {
        "project":{"key": "T2L"},
        "summary": df.loc[first_index,'Summary'],
        "assignee": {"name": df.loc[first_index,'Assignee']},
        "reporter": {"name": df.loc[first_index,'Reporter']},
        "priority": {"name": df.loc[first_index,'Priority']},
        # Test Repository Path: customfield_17291
        "customfield_17291": df.loc[first_index,'Location'],
        "description": df.loc[first_index,'Description'],
        # "customfield_15380" "responsible"
        "customfield_15380": {"name": df.loc[first_index,'Responsible']},
        # "Affected Application/s" "customfield_16181"
        "customfield_16181": [df.loc[first_index,'Affected Application/s']],
        "issuetype": {"name": "Test"},
        # Manual test steps: customfield_17284
        "customfield_17284": wrapper_of_wrapper
        }
        return payload
    else:
        logging.getLogger(__name__).info(f"File is empty: {file}")
        
def create_ticket(payload):
    loader = Loader()

    jira = atlassian.Jira(url = "https://jira/", username = loader.cfg.user, password = loader.cfg.password)
    response = jira.issue_create(fields=payload)
    key = response["key"]
    jira.issue_transition(key, "In Use")
    logging.getLogger(__name__).info(f"{key} created")
    return key

def ticket_does_not_exist(path_to_mapping, cur_file):
    mapping = pd.read_excel(path_to_mapping, sheet_name="Reports")
    key = mapping['JIRA_Key'][mapping["Filename"] == cur_file]
    
    if key.values[0] is nan:
        logging.getLogger(__name__).info(f"{cur_file} not found, calling JIRA webservice for creation")
        return True
    elif key.values[0] is not nan and isinstance(key.values[0],str):
        if len(key.values[0]) > 0:
            logging.getLogger(__name__).info(f"{cur_file} found on the mapping file, cancel creation")
            return False
    return True

def setLogger() -> logging.Logger:
    ''' check if FileHandler exists'''
    ''' FileHandler is created at python level and not @ the level of this class '''
    if (len(logging.getLogger(__name__).handlers) == 0):
        ''' Instantiate the FileHandler aka logger object so that it records general info and exceptions into a .log file '''
        logger = logging.getLogger(__name__)
        logger.setLevel(logging.INFO)
        ''' create a file handler '''
        handler = logging.FileHandler('aris.log')
        handler.setLevel(logging.INFO)
        ''' create a logging format '''
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        handler.setFormatter(formatter)
        ''' add the handlers to the logger '''
        logger.addHandler(handler)
        return logger

def main():
    """
    This script checks creates new test cases based on the mapping file. If there is no JIRA key attached to the test case and the process justifies 
    the creation of a test case then a ticket is created. 
    """
    setLogger()

#    files = [".xls"]
    
    keep_track_test_case_num = 1
    parent = os.getcwd()
    path_to_mapping = os.path.join(parent, "ARIS", "mapping.xlsx")
    payloads, data = list(), dict()
    
    for file in files:
        logging.getLogger(__name__).info(f"Processing started for: {file}")
        path_to_file = os.path.join(parent, "ARIS", "aris export", file)
        process = pd.read_excel(path_to_file, encoding='iso-8859-1', sheet_name="Table 2", header=None)
        business_rep = pd.read_excel(path_to_file, encoding='iso-8859-1', sheet_name="Table 1", header=None)
        final_json = mash_data(process)
        df = convert_to_df(final_json)
        df = complement_missing_data(df)
        ordering = grab_first_item(df)
        df = apply_ordering(df, ordering)
        df = convert_to_xray_format(df, file, keep_track_test_case_num, business_rep)
        df = add_url(df, path_to_mapping, file)
        payload = create_ticket_payload(df, file)
        if ticket_does_not_exist(path_to_mapping, file) and df.empty is False:
            _id = payload['description']
            data[_id] = payload
            jira_key = create_ticket(payload)
            add_jira_key(jira_key, path_to_mapping, file)
            payloads.append(data)
            generate_excel(df, file)
            keep_track_test_case_num += 1
        elif df.empty is True:
            logging.getLogger(__name__).info(f"Webservice not called as there is nothing to create. The process on ARIS does not justify the test case creation")

    timerz = time.strftime("_%Y%m%d_%H%M%S")
    if keep_track_test_case_num-1 == 1:
        logging.getLogger(__name__).info(f"{keep_track_test_case_num-1} test case was created")
    logging.getLogger(__name__).info(f"{keep_track_test_case_num-1} test cases were created")
    outfile_name = 'data' + timerz + '.json'
    with open(outfile_name, 'w') as outfile:
        json.dump(payloads, outfile)
    consolidated = compress_into_one_file()
    generate_csv(consolidated, "Consolidated")
    generate_excel(consolidated, "Consolidated")

if __name__ == "__main__":
    main()


